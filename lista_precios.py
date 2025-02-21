import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

def parse_price(price_str):
    """Convierte un string con formato '$19.400' a valor numérico."""
    price_str = price_str.replace("$", "").strip()
    price_str = price_str.replace(".", "")
    try:
        return float(price_str)
    except Exception as e:
        print(f"Error al parsear el precio '{price_str}': {e}")
        return 0.0

def convertir_a_gramos(label):
    """Convierte '100g' o '1kg' a su valor en gramos."""
    label = label.lower().strip()
    if label.endswith("kg"):
        try:
            return float(label.replace("kg", "").strip()) * 1000
        except:
            return None
    elif label.endswith("g"):
        try:
            return float(label.replace("g", "").strip())
        except:
            return None
    else:
        return None

def calcular_precio(frac_label, precio_base):
    """Calcula el precio para un fraccionamiento dado el precio base (por kg)."""
    gramos = convertir_a_gramos(frac_label)
    if gramos is None:
        return None
    return int(round((gramos / 1000) * precio_base))

def format_price(num):
    """
    Convierte un entero (p.ej. 2079) en un string con separador de miles '.' y signo '$'.
    Ej: 2079 -> "$2.079"
    """
    s = f"{num:,}"  # "2,079"
    s = s.replace(",", ".")  # "2.079"
    return f"${s}"

def generar_lista_precios(csv_path, encoding='utf-8'):
    """
    Lee el CSV con la codificación especificada (por defecto UTF-8) y
    construye un DataFrame con filas de categoría y producto.
    """
    df = pd.read_csv(csv_path, encoding=encoding)  # <--- AJUSTA AQUÍ la codificación real
    required_cols = ["CATEGORIA", "PRODUCTO", "KG / UNIDAD", "PRECIO VENTA", "FRACCIONAMIENTO", "MARCA"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna {col} en el CSV.")

    output_rows = []
    row_types = []  # Para indicar si la fila es 'category' o 'product'
    
    grouped = df.groupby("CATEGORIA", sort=False)
    for cat, group in grouped:
        tipos = group["KG / UNIDAD"].astype(str).str.strip().str.upper().unique().tolist()
        is_mixed = ("KG" in tipos) and ("UNIDAD" in tipos)
        is_pure_kg = (len(tipos) == 1 and tipos[0] == "KG")
        is_pure_unidad = (len(tipos) == 1 and tipos[0] == "UNIDAD")
        
        # Reunimos los fraccionamientos para la categoría
        kg_fracs = []
        for _, row in group.iterrows():
            tipo = str(row["KG / UNIDAD"]).strip().upper()
            if tipo == "KG":
                frac_raw = row["FRACCIONAMIENTO"]
                frac_str = str(frac_raw).strip() if pd.notna(frac_raw) else ""
                if frac_str:
                    for frac in frac_str.split(","):
                        frac = frac.strip()
                        if frac and frac not in kg_fracs:
                            kg_fracs.append(frac)
        kg_fracs_sorted = sorted(
            kg_fracs,
            key=lambda x: convertir_a_gramos(x) if convertir_a_gramos(x) else 0
        )
        
        allowed_kg = 2 if is_mixed else 3
        if len(kg_fracs_sorted) > allowed_kg:
            print(f"ADVERTENCIA: En la categoría '{cat}' se encontraron más de {allowed_kg} fraccionamientos: {kg_fracs_sorted}. Se usarán los primeros {allowed_kg}.")
            kg_fracs_sorted = kg_fracs_sorted[:allowed_kg]

        # Fila de categoría
        header = [cat, "", "", "", "MARCA"]
        if is_pure_unidad:
            header[3] = "UNIDAD"
        elif is_pure_kg:
            if len(kg_fracs_sorted) == 1:
                header[3] = kg_fracs_sorted[0]
            elif len(kg_fracs_sorted) == 2:
                header[2] = kg_fracs_sorted[0]
                header[3] = kg_fracs_sorted[1]
            elif len(kg_fracs_sorted) >= 3:
                header[1] = kg_fracs_sorted[0]
                header[2] = kg_fracs_sorted[1]
                header[3] = kg_fracs_sorted[2]
        elif is_mixed:
            if len(kg_fracs_sorted) == 1:
                header[2] = kg_fracs_sorted[0]
            elif len(kg_fracs_sorted) == 2:
                header[1] = kg_fracs_sorted[0]
                header[2] = kg_fracs_sorted[1]
            header[3] = "UNIDAD"
        else:
            header[3] = ""
        
        output_rows.append(header)
        row_types.append("category")
        
        # Filas de producto
        for _, row in group.iterrows():
            prod = row["PRODUCTO"]
            tipo = str(row["KG / UNIDAD"]).strip().upper()
            precio_base = parse_price(str(row["PRECIO VENTA"]))
            marca = str(row["MARCA"]).strip() if pd.notna(row["MARCA"]) else ""
            
            prod_row = [prod, "", "", "", marca]
            
            if tipo == "UNIDAD":
                price_int = int(round(precio_base))
                prod_row[3] = format_price(price_int)
            
            elif tipo == "KG":
                frac_raw = row["FRACCIONAMIENTO"]
                frac_str = str(frac_raw).strip() if pd.notna(frac_raw) else ""
                if not frac_str:
                    print(f"ADVERTENCIA: El producto '{prod}' (categoría '{cat}') no tiene fraccionamiento.")
                else:
                    fracs = [x.strip() for x in frac_str.split(",") if x.strip()]
                    if is_mixed and len(fracs) > 2:
                        print(f"ADVERTENCIA: El producto '{prod}' (categoría '{cat}') tiene más de 2 fraccionamientos: {fracs}.")
                        fracs = fracs[:2]
                    elif is_pure_kg and len(fracs) > 3:
                        print(f"ADVERTENCIA: El producto '{prod}' (categoría '{cat}') tiene más de 3 fraccionamientos: {fracs}.")
                        fracs = fracs[:3]
                    fracs_sorted = sorted(
                        fracs,
                        key=lambda x: convertir_a_gramos(x) if convertir_a_gramos(x) else 0
                    )
                    
                    precios = []
                    for frac in fracs_sorted:
                        p = calcular_precio(frac, precio_base)
                        if p is not None:
                            precios.append(format_price(p))
                        else:
                            precios.append("")
                    
                    if is_mixed:
                        if len(precios) == 1:
                            prod_row[2] = precios[0]  # C
                        elif len(precios) == 2:
                            prod_row[1] = precios[0]  # B
                            prod_row[2] = precios[1]  # C
                    else:
                        if len(precios) == 1:
                            prod_row[3] = precios[0]
                        elif len(precios) == 2:
                            prod_row[2] = precios[0]
                            prod_row[3] = precios[1]
                        elif len(precios) >= 3:
                            prod_row[1] = precios[0]
                            prod_row[2] = precios[1]
                            prod_row[3] = precios[2]
            
            output_rows.append(prod_row)
            row_types.append("product")
    
    df_out = pd.DataFrame(output_rows, columns=["CATEGORIA / PRODUCTO", "COL B", "COL C", "COL D", "COL E"])
    return df_out, row_types


if __name__ == "__main__":
    from openpyxl.styles import Alignment, Font, PatternFill
    # 1) Estilos para las filas 1..4
    # Fila 1: color de fondo #cccccc, texto negro, Arial 15, negrita
    fill_row1 = PatternFill("solid", fgColor="FFcccccc")  
    font_row1 = Font(name="Arial", size=15, bold=True, color="FF000000")  
    
    # Fila 2: Oswald 10, color texto negro, fondo #ff9900
    fill_row2 = PatternFill("solid", fgColor="FFFF9900")
    font_row2 = Font(name="Oswald", size=10, bold=False, color="FF000000")
    
    # Fila 3: Oswald 9, negrita, color texto negro, fondo #ffe599
    fill_row3 = PatternFill("solid", fgColor="FFffe599")
    font_row3 = Font(name="Oswald", size=9, bold=True, color="FF000000")
    
    # Fila 4: Oswald 10, negrita, color texto #2c2c4d, fondo #b4a7d6
    fill_row4 = PatternFill("solid", fgColor="FFb4a7d6")
    font_row4 = Font(name="Oswald", size=10, bold=True, color="FF2c2c4d")
    
    # 2) Estilos para categorías y productos
    fill_green  = PatternFill("solid", fgColor="FF93C47D")  # Verde para categoría
    font_category = Font(name="Oswald", size=11, bold=True, color="FF000000")  # Categoría
    font_product  = Font(name="Candara", size=11, color="FF595959")            # Texto gris en productos
    font_price    = Font(name="Candara", size=12, color="FF595959")            # Precios en productos
    
    # === Generar el DataFrame con la codificación adecuada ===
    archivo_csv = "LISTA PRECIOS - PRODUCTOS (6).csv"
    df_resultado, row_types = generar_lista_precios(archivo_csv, encoding="utf-8")  # <--- AJUSTA encoding según tu CSV

    with pd.ExcelWriter("estructura_listaprecios.xlsx", engine="openpyxl") as writer:
        # Insertamos el DataFrame en la hoja "Sheet1" a partir de la fila 5
        df_resultado.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=4,
            header=False,
            index=False
        )
        ws = writer.sheets["Sheet1"]
        
        # Ajuste de anchos de columna
        ws.column_dimensions["A"].width = 92.55
        ws.column_dimensions["B"].width = 14.82
        ws.column_dimensions["C"].width = 14.82
        ws.column_dimensions["D"].width = 14.82
        ws.column_dimensions["E"].width = 20
        
        # === Aplicamos estilo a las filas 1..4 ===
        # Fila 1
        ws["A1"] = "LISTA DE PRECIOS FEBRERO 2025"
        for col in range(1, 6):
            c = ws.cell(row=1, column=col)
            c.font = font_row1
            c.fill = fill_row1
        
        # Fila 2
        ws["A2"] = "ENVIOS SIN CARGO ( SEGÚN MONTO PREESTABLECIDO PREVIAMENTE A LA ENTREGA )"
        for col in range(1, 6):
            c = ws.cell(row=2, column=col)
            c.font = font_row2
            c.fill = fill_row2
        
        # Fila 3
        ws["A3"] = "PRECIOS SUJETOS A CAMBIOS SIN PREVIO AVISO . SIEMPRE TRATAREMOS DE INFORMARLOS PREVIAMENTE"
        for col in range(1, 6):
            c = ws.cell(row=3, column=col)
            c.font = font_row3
            c.fill = fill_row3
        
        # Fila 4
        ws["A4"] = "MEDIOS DE PAGO : EFECTIVO / TRANSFERENCIA BANCARIA O DIGITAL / DEBITO"
        for col in range(1, 6):
            c = ws.cell(row=4, column=col)
            c.font = font_row4
            c.fill = fill_row4
        
        # === Aplicamos estilo a las filas de la tabla (desde la fila 5 en adelante) ===
        for i, rtype in enumerate(row_types):
            excel_row = i + 5  # La primera fila de la tabla es la 5
            if rtype == "category":
                # Fondo verde y fuente Oswald 11
                for col in range(1, 6):
                    cell = ws.cell(row=excel_row, column=col)
                    cell.fill = fill_green
                    cell.font = font_category
            else:
                # Fila de producto => sin relleno, texto gris
                ws.cell(row=excel_row, column=1).font = font_product
                ws.cell(row=excel_row, column=5).font = font_product
                for col in [2, 3, 4]:
                    c = ws.cell(row=excel_row, column=col)
                    if c.value:
                        c.font = font_price
                    else:
                        c.font = font_product

        for row_idx in range(5, ws.max_row + 1):
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        