# lista_precios_utils.py
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill

# Importamos las funciones de cálculo unificado (asegúrate de tener este módulo implementado)
from price_calculator import compute_fraction_price, convertir_a_gramos

def format_price(num):
    """
    Convierte un entero (p.ej. 2079) en un string con separador de miles '.' y signo '$'.
    Ejemplo: 2079 -> "$2.079"
    """
    s = f"{num:,}"
    s = s.replace(",", ".")
    return f"${s}"

def generar_lista_precios_df(df):
    """
    Genera un DataFrame con la lista de precios a partir del DataFrame de productos.
    Se espera que el DataFrame contenga las columnas:
    "CATEGORIA", "PRODUCTO", "KG / UNIDAD", "PRECIO VENTA", "FRACCIONAMIENTO", "MARCA" y "STOCK".
    Retorna un DataFrame resultante y una lista que indica el tipo de cada fila ("category" o "product").
    """
    required_cols = ["CATEGORIA", "PRODUCTO", "KG / UNIDAD", "PRECIO VENTA", "FRACCIONAMIENTO", "MARCA", "STOCK"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna {col} en el DataFrame.")
    
    output_rows = []
    row_types = []  # Indica si la fila es 'category' o 'product'
    
    grouped = df.groupby("CATEGORIA", sort=False)
    for cat, group in grouped:
        tipos = group["KG / UNIDAD"].astype(str).str.strip().str.upper().unique().tolist()
        is_mixed = ("KG" in tipos) and ("UNIDAD" in tipos)
        is_pure_kg = (len(tipos) == 1 and tipos[0] == "KG")
        is_pure_unidad = (len(tipos) == 1 and tipos[0] == "UNIDAD")
        
        # Reunir fraccionamientos para productos vendidos por KG en la categoría
        kg_fracs = []
        for _, row in group.iterrows():
            tipo = str(row["KG / UNIDAD"]).strip().upper()
            if tipo == "KG":
                frac_raw = row["FRACCIONAMIENTO"]
                frac_str = str(frac_raw).strip() if pd.notna(frac_raw) else ""
                if frac_str:
                    for frac in frac_str.split(","):
                        frac = frac.strip()
                        if frac and frac not in kg_fracs:
                            kg_fracs.append(frac)
        kg_fracs_sorted = sorted(
            kg_fracs,
            key=lambda x: convertir_a_gramos(x) if convertir_a_gramos(x) is not None else 0
        )
        
        allowed_kg = 2 if is_mixed else 3
        if len(kg_fracs_sorted) > allowed_kg:
            print(f"ADVERTENCIA: En la categoría '{cat}' se encontraron más de {allowed_kg} fraccionamientos: {kg_fracs_sorted}. Se usarán los primeros {allowed_kg}.")
            kg_fracs_sorted = kg_fracs_sorted[:allowed_kg]

        # Fila de categoría (cabecera)
        header = [cat, "", "", "", "MARCA"]
        if is_pure_unidad:
            header[3] = "UNIDAD"
        elif is_pure_kg:
            if len(kg_fracs_sorted) == 1:
                header[3] = kg_fracs_sorted[0]
            elif len(kg_fracs_sorted) == 2:
                header[2] = kg_fracs_sorted[0]
                header[3] = kg_fracs_sorted[1]
            elif len(kg_fracs_sorted) >= 3:
                header[1] = kg_fracs_sorted[0]
                header[2] = kg_fracs_sorted[1]
                header[3] = kg_fracs_sorted[2]
        elif is_mixed:
            if len(kg_fracs_sorted) == 1:
                header[2] = kg_fracs_sorted[0]
            elif len(kg_fracs_sorted) == 2:
                header[1] = kg_fracs_sorted[0]
                header[2] = kg_fracs_sorted[1]
            header[3] = "UNIDAD"
        else:
            header[3] = ""
        
        output_rows.append(header)
        row_types.append("category")
        
        # Filas para cada producto
        for _, row in group.iterrows():
            prod = row["PRODUCTO"]
            tipo = str(row["KG / UNIDAD"]).strip().upper()
            try:
                precio_base = float(row["PRECIO VENTA"])
            except:
                precio_base = 0.0
            marca = str(row["MARCA"]).strip() if pd.notna(row["MARCA"]) else ""
            
            # Iniciamos la fila con: [Producto, COL B, COL C, COL D, Marca]
            prod_row = [prod, "", "", "", marca]
            
            # Verificar el stock: '-' indica stock disponible; "0" indica sin stock.
            stock_val = str(row["STOCK"]).strip()
            if stock_val == "0":
                # Producto sin stock: se reemplazan las casillas de precio por "SIN STOCK"
                if tipo == "UNIDAD":
                    prod_row[3] = "SIN STOCK"
                elif tipo == "KG":
                    # Para KG, independientemente de la cantidad de fraccionamientos esperados,
                    # rellenamos las columnas destinadas a precios (indices 1,2 y 3) con "SIN STOCK"
                    prod_row[1] = "SIN STOCK"
                    prod_row[2] = "SIN STOCK"
                    prod_row[3] = "SIN STOCK"
            else:
                # Si hay stock, se procede a calcular los precios.
                if tipo == "UNIDAD":
                    prod_row[3] = format_price(int(round(precio_base)))
                elif tipo == "KG":
                    frac_raw = row["FRACCIONAMIENTO"]
                    frac_str = str(frac_raw).strip() if pd.notna(frac_raw) else ""
                    if not frac_str:
                        print(f"ADVERTENCIA: El producto '{prod}' (categoría '{cat}') no tiene fraccionamiento.")
                    else:
                        fracs = [x.strip() for x in frac_str.split(",") if x.strip()]
                        if is_mixed and len(fracs) > 2:
                            print(f"ADVERTENCIA: El producto '{prod}' (categoría '{cat}') tiene más de 2 fraccionamientos: {fracs}.")
                            fracs = fracs[:2]
                        elif is_pure_kg and len(fracs) > 3:
                            print(f"ADVERTENCIA: El producto '{prod}' (categoría '{cat}') tiene más de 3 fraccionamientos: {fracs}.")
                            fracs = fracs[:3]
                        fracs_sorted = sorted(
                            fracs,
                            key=lambda x: convertir_a_gramos(x) if convertir_a_gramos(x) is not None else 0
                        )
                        
                        precios = []
                        for frac in fracs_sorted:
                            p = compute_fraction_price(tipo, precio_base, frac)
                            if p is not None:
                                precios.append(format_price(int(round(p))))
                            else:
                                precios.append("")
                        
                        if is_mixed:
                            if len(precios) == 1:
                                prod_row[2] = precios[0]
                            elif len(precios) == 2:
                                prod_row[1] = precios[0]
                                prod_row[2] = precios[1]
                        else:
                            if len(precios) == 1:
                                prod_row[3] = precios[0]
                            elif len(precios) == 2:
                                prod_row[2] = precios[0]
                                prod_row[3] = precios[1]
                            elif len(precios) >= 3:
                                prod_row[1] = precios[0]
                                prod_row[2] = precios[1]
                                prod_row[3] = precios[2]
            
            output_rows.append(prod_row)
            row_types.append("product")
    
    df_out = pd.DataFrame(output_rows, columns=["CATEGORIA / PRODUCTO", "COL B", "COL C", "COL D", "COL E"])
    return df_out, row_types

# Definición de estilos para el Excel
fill_row1 = PatternFill("solid", fgColor="FFcccccc")  
font_row1 = Font(name="Arial", size=15, bold=True, color="FF000000")  

fill_row2 = PatternFill("solid", fgColor="FFFF9900")
font_row2 = Font(name="Oswald", size=10, bold=False, color="FF000000")

fill_row3 = PatternFill("solid", fgColor="FFffe599")
font_row3 = Font(name="Oswald", size=9, bold=True, color="FF000000")

fill_row4 = PatternFill("solid", fgColor="FFb4a7d6")
font_row4 = Font(name="Oswald", size=10, bold=True, color="FF2c2c4d")

fill_green  = PatternFill("solid", fgColor="FF93C47D")  # Verde para categorías
font_category = Font(name="Oswald", size=11, bold=True, color="FF000000")  
font_product  = Font(name="Candara", size=11, color="FF595959")            
font_price    = Font(name="Candara", size=12, color="FF595959")            

def crear_excel_con_estilo(df_out, row_types, title="LISTA DE PRECIOS FEBRERO 2025"):
    """
    Crea un archivo Excel en memoria con el DataFrame generado y
    aplica estilos a las primeras filas y a la tabla.
    Retorna un objeto BytesIO con el contenido del Excel.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Escribimos el DataFrame a partir de la fila 5 (índice 4)
        df_out.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=4,
            header=False,
            index=False
        )
        ws = writer.sheets["Sheet1"]
        
        # Ajuste de anchos de columna
        ws.column_dimensions["A"].width = 92.55
        ws.column_dimensions["B"].width = 14.82
        ws.column_dimensions["C"].width = 14.82
        ws.column_dimensions["D"].width = 14.82
        ws.column_dimensions["E"].width = 20
        
        # Aplicar estilos a las filas 1 a 4
        ws["A1"] = title
        for col in range(1, 6):
            c = ws.cell(row=1, column=col)
            c.font = font_row1
            c.fill = fill_row1
        
        ws["A2"] = "ENVIOS SIN CARGO ( SEGÚN MONTO PREESTABLECIDO PREVIAMENTE A LA ENTREGA )"
        for col in range(1, 6):
            c = ws.cell(row=2, column=col)
            c.font = font_row2
            c.fill = fill_row2
        
        ws["A3"] = "PRECIOS SUJETOS A CAMBIOS SIN PREVIO AVISO . SIEMPRE TRATAREMOS DE INFORMARLOS PREVIAMENTE"
        for col in range(1, 6):
            c = ws.cell(row=3, column=col)
            c.font = font_row3
            c.fill = fill_row3
        
        ws["A4"] = "MEDIOS DE PAGO : EFECTIVO / TRANSFERENCIA BANCARIA O DIGITAL / DEBITO"
        for col in range(1, 6):
            c = ws.cell(row=4, column=col)
            c.font = font_row4
            c.fill = fill_row4
        
        # Aplicar estilo a las filas de la tabla (desde la fila 5 en adelante)
        for i, rtype in enumerate(row_types):
            excel_row = i + 5  # La primera fila de la tabla es la 5
            if rtype == "category":
                for col in range(1, 6):
                    cell = ws.cell(row=excel_row, column=col)
                    cell.fill = fill_green
                    cell.font = font_category
            else:
                ws.cell(row=excel_row, column=1).font = font_product
                ws.cell(row=excel_row, column=5).font = font_product
                for col in [2, 3, 4]:
                    c = ws.cell(row=excel_row, column=col)
                    if c.value:
                        c.font = font_price
                    else:
                        c.font = font_product
        
        # Configurar la alineación de todas las celdas de la tabla
        for row_idx in range(5, ws.max_row + 1):
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    output.seek(0)
    return output
