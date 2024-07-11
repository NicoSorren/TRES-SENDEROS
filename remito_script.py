import shutil
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import datetime
import os

# Diccionario de categorías (mantenlo como está)
categoria_columna_precio = {
    '001': {'nombre': 0, 'precio_250gr': 2},
    '035': {'nombre': 0, 'unidad': 3},
    '037': {'nombre': 0, 'unidad': 3},
    # ... (otros elementos del diccionario)
}

def cargar_productos():
    df = pd.read_excel('LISTA DE PRECIOS JUNIO TS-2 (1).xlsx', header=None)
    productos = []
    codigo_categoria = None

    for i in range(len(df)):
        if pd.isnull(df.iloc[i, 5]):
            continue

        codigo = str(int(df.iloc[i, 5])).strip().zfill(3)

        if codigo in categoria_columna_precio:
            codigo_categoria = codigo
            categoria_info = categoria_columna_precio[codigo_categoria]
            productos.append({"nombre": f"Categoría: {codigo_categoria}", "codigo": codigo_categoria, **categoria_info})
            continue

        if codigo_categoria and codigo.isdigit():
            columnas = categoria_columna_precio[codigo_categoria]
            producto_nombre = {"nombre": df.iloc[i, columnas['nombre']]}
            for clave in ['precio_100gr', 'precio_250gr', 'precio_500gr', '1kg', 'unidad']:
                if clave in columnas:
                    producto_nombre[clave] = df.iloc[i, columnas[clave]]
            producto_nombre["codigo"] = codigo
            productos.append(producto_nombre)
    return productos

def eliminar_imagen_por_id(archivo_excel, id_imagen):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active

    for img in ws._images:
        if img.path and os.path.basename(img.path) == id_imagen:
            ws._images.remove(img)
            break
    wb.save(archivo_excel)

def escribir_remito(nombre, fecha, direccion, telefono, localidad, observaciones, seleccionados, subtotal, descuento, total_precio):
    archivo_excel_original = 'FORMATO REMITO.xlsx'  # Ruta relativa
    archivo_excel = f'TEMP_{archivo_excel_original}'

    shutil.copyfile(archivo_excel_original, archivo_excel)
    eliminar_imagen_por_id(archivo_excel, 'image1.png')

    wb = load_workbook(archivo_excel)
    ws = wb.active
    ws['B9'] = nombre
    ws['D9'] = fecha
    ws['B10'] = direccion
    ws['B11'] = telefono
    ws['B12'] = localidad
    ws['B15'] = observaciones

    start_row = 18
    for i, item in enumerate(seleccionados):
        row = start_row + i
        ws[f'A{row}'] = f"{item['peso']}g"
        ws[f'B{row}'] = item['producto']
        ws[f'C{row}'] = f"${item['precio_por_100gr']:.2f}"
        ws[f'D{row}'] = f"${item['precio_total']:.2f}"

    ws['D41'] = f"${subtotal:.2f}"
    descuento_total = subtotal * (descuento / 100)
    ws['D42'] = f"${descuento_total:.2f}"
    ws['D57'] = f"${total_precio:.2f}"

    nombre_archivo_excel = f"remito_{nombre}_{fecha}.xlsx"

    try:
        wb.save(nombre_archivo_excel)
    except Exception as e:
        print(f"An error occurred: {e}")

    print(f"Remito generado: {nombre_archivo_excel}")
    return nombre_archivo_excel
